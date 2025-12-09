import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import mysql.connector
from datetime import datetime, timedelta
import hashlib
from fpdf import FPDF
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class Database:
    def __init__(self):
        try:
            temp_conn = mysql.connector.connect(
                host="localhost",
                user="root",
                password="root"
            )
            temp_cursor = temp_conn.cursor()
            temp_cursor.execute("CREATE DATABASE IF NOT EXISTS expense_tracker")
            temp_conn.close()
        except mysql.connector.Error as err:
            messagebox.showerror("Database Error", f"MySQL connection failed!\n\nError: {err}")
            raise
        
        try:
            self.conn = mysql.connector.connect(
                host="localhost",
                user="root",
                password="root",
                database="expense_tracker"
            )
            self.cursor = self.conn.cursor()
            self.create_tables()
        except mysql.connector.Error as err:
            messagebox.showerror("Database Error", f"Failed to connect: {err}")
            raise
    
    def create_tables(self):
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INT AUTO_INCREMENT PRIMARY KEY,
                username VARCHAR(100) UNIQUE NOT NULL,
                email VARCHAR(100) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                monthly_budget DECIMAL(10,2) DEFAULT 0,
                dark_mode BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                is_active BOOLEAN DEFAULT TRUE
            )
        """)
        
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS expenses (
                id INT AUTO_INCREMENT PRIMARY KEY,
                user_id INT,
                expense_name VARCHAR(255),
                category VARCHAR(100) DEFAULT 'Other',
                amount DECIMAL(10,2),
                exp_date DATE,
                exp_time TIME,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users(id)
            )
        """)
        
        admin_pass = hashlib.sha256("admin123".encode()).hexdigest()
        try:
            self.cursor.execute(
                "INSERT INTO users (username, email, password) VALUES (%s, %s, %s)",
                ("admin", "admin@expense.com", admin_pass)
            )
            self.conn.commit()
        except:
            pass
    
    def register_user(self, username, email, password):
        hashed_pw = hashlib.sha256(password.encode()).hexdigest()
        try:
            self.cursor.execute(
                "INSERT INTO users (username, email, password) VALUES (%s, %s, %s)",
                (username, email, hashed_pw)
            )
            self.conn.commit()
            return True
        except mysql.connector.IntegrityError:
            return False
    
    def login_user(self, username, password):
        hashed_pw = hashlib.sha256(password.encode()).hexdigest()
        self.cursor.execute(
            "SELECT id, username, monthly_budget, dark_mode FROM users WHERE username=%s AND password=%s",
            (username, hashed_pw)
        )
        return self.cursor.fetchone()
    
    def add_expense(self, user_id, name, category, amount, date, time):
        self.cursor.execute(
            "INSERT INTO expenses (user_id, expense_name, category, amount, exp_date, exp_time) VALUES (%s, %s, %s, %s, %s, %s)",
            (user_id, name, category, amount, date, time)
        )
        self.conn.commit()
    
    def update_expense(self, exp_id, name, category, amount, date, time):
        self.cursor.execute(
            "UPDATE expenses SET expense_name=%s, category=%s, amount=%s, exp_date=%s, exp_time=%s WHERE id=%s",
            (name, category, amount, date, time, exp_id)
        )
        self.conn.commit()
    
    def delete_expense(self, exp_id):
        self.cursor.execute("DELETE FROM expenses WHERE id=%s", (exp_id,))
        self.conn.commit()
    
    def get_expenses(self, user_id, start_date=None, end_date=None):
        if start_date and end_date:
            self.cursor.execute(
                "SELECT id, expense_name, category, amount, exp_date, exp_time FROM expenses WHERE user_id=%s AND exp_date BETWEEN %s AND %s ORDER BY exp_date DESC",
                (user_id, start_date, end_date)
            )
        else:
            self.cursor.execute(
                "SELECT id, expense_name, category, amount, exp_date, exp_time FROM expenses WHERE user_id=%s ORDER BY exp_date DESC",
                (user_id,)
            )
        return self.cursor.fetchall()
    
    def get_total_expense(self, user_id, start_date=None, end_date=None):
        if start_date and end_date:
            self.cursor.execute(
                "SELECT SUM(amount) FROM expenses WHERE user_id=%s AND exp_date BETWEEN %s AND %s",
                (user_id, start_date, end_date)
            )
        else:
            self.cursor.execute(
                "SELECT SUM(amount) FROM expenses WHERE user_id=%s",
                (user_id,)
            )
        result = self.cursor.fetchone()[0]
        return result if result else 0
    
    def get_category_totals(self, user_id, start_date=None, end_date=None):
        if start_date and end_date:
            self.cursor.execute(
                "SELECT category, SUM(amount) FROM expenses WHERE user_id=%s AND exp_date BETWEEN %s AND %s GROUP BY category",
                (user_id, start_date, end_date)
            )
        else:
            self.cursor.execute(
                "SELECT category, SUM(amount) FROM expenses WHERE user_id=%s GROUP BY category",
                (user_id,)
            )
        return self.cursor.fetchall()
    
    def update_budget(self, user_id, budget):
        self.cursor.execute("UPDATE users SET monthly_budget=%s WHERE id=%s", (budget, user_id))
        self.conn.commit()
    
    def toggle_dark_mode(self, user_id, mode):
        self.cursor.execute("UPDATE users SET dark_mode=%s WHERE id=%s", (mode, user_id))
        self.conn.commit()
    
    def get_all_users(self):
        self.cursor.execute(
            "SELECT username, email, created_at, is_active FROM users WHERE username != 'admin'"
        )
        return self.cursor.fetchall()

class ExpenseTrackerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Advanced Expense Tracker Pro")
        self.root.geometry("1200x750")
        
        self.CATEGORIES = ["Food & Dining", "Transportation", "Shopping", "Healthcare", 
                          "Entertainment", "Bills & Utilities", "Education", "Other"]
        
        try:
            self.db = Database()
            self.current_user = None
            self.dark_mode = False
            self.show_login()
        except Exception as e:
            messagebox.showerror("Startup Error", "Failed to initialize application!")
            self.root.destroy()
    
    def apply_theme(self):
        if self.dark_mode:
            self.bg_color = "#1e1e1e"
            self.fg_color = "#ffffff"
            self.frame_bg = "#2d2d2d"
            self.button_bg = "#0d7377"
            self.highlight = "#14ffec"
        else:
            self.bg_color = "#f0f0f0"
            self.fg_color = "#000000"
            self.frame_bg = "#ffffff"
            self.button_bg = "#4CAF50"
            self.highlight = "#2196F3"
        
        self.root.configure(bg=self.bg_color)
    
    def clear_window(self):
        for widget in self.root.winfo_children():
            widget.destroy()
    
    def show_login(self):
        self.clear_window()
        self.root.configure(bg="#f0f0f0")
        
        frame = tk.Frame(self.root, bg="#ffffff", relief="raised", bd=2)
        frame.place(relx=0.5, rely=0.5, anchor="center")
        
        tk.Label(frame, text="🔐 Login", font=("Arial", 28, "bold"), bg="#ffffff", fg="#2196F3").grid(row=0, column=0, columnspan=2, pady=25, padx=40)
        
        tk.Label(frame, text="Username:", font=("Arial", 12), bg="#ffffff").grid(row=1, column=0, sticky="e", padx=10, pady=10)
        username_entry = tk.Entry(frame, font=("Arial", 12), width=25, relief="solid", bd=1)
        username_entry.grid(row=1, column=1, padx=10, pady=10)
        
        tk.Label(frame, text="Password:", font=("Arial", 12), bg="#ffffff").grid(row=2, column=0, sticky="e", padx=10, pady=10)
        password_entry = tk.Entry(frame, font=("Arial", 12), width=25, show="●", relief="solid", bd=1)
        password_entry.grid(row=2, column=1, padx=10, pady=10)
        
        def login():
            user = self.db.login_user(username_entry.get(), password_entry.get())
            if user:
                self.current_user = {"id": user[0], "username": user[1], "budget": user[2]}
                self.dark_mode = bool(user[3])
                if user[1] == "admin":
                    self.show_admin_dashboard()
                else:
                    self.show_home()
            else:
                messagebox.showerror("Error", "Invalid credentials!")
        
        tk.Button(frame, text="Login", font=("Arial", 12, "bold"), bg="#4CAF50", fg="white", 
                 width=15, height=2, command=login, cursor="hand2").grid(row=3, column=0, columnspan=2, pady=20)
        
        tk.Button(frame, text="Register New User", font=("Arial", 10), bg="#2196F3", fg="white",
                 command=self.show_register, cursor="hand2").grid(row=4, column=0, columnspan=2, pady=(0, 20))
    
    def show_register(self):
        self.clear_window()
        self.root.configure(bg="#f0f0f0")
        
        frame = tk.Frame(self.root, bg="#ffffff", relief="raised", bd=2)
        frame.place(relx=0.5, rely=0.5, anchor="center")
        
        tk.Label(frame, text="📝 Registration", font=("Arial", 28, "bold"), bg="#ffffff", fg="#2196F3").grid(row=0, column=0, columnspan=2, pady=25, padx=40)
        
        tk.Label(frame, text="Username:", font=("Arial", 12), bg="#ffffff").grid(row=1, column=0, sticky="e", padx=10, pady=10)
        username_entry = tk.Entry(frame, font=("Arial", 12), width=25, relief="solid", bd=1)
        username_entry.grid(row=1, column=1, padx=10, pady=10)
        
        tk.Label(frame, text="Email:", font=("Arial", 12), bg="#ffffff").grid(row=2, column=0, sticky="e", padx=10, pady=10)
        email_entry = tk.Entry(frame, font=("Arial", 12), width=25, relief="solid", bd=1)
        email_entry.grid(row=2, column=1, padx=10, pady=10)
        
        tk.Label(frame, text="Password:", font=("Arial", 12), bg="#ffffff").grid(row=3, column=0, sticky="e", padx=10, pady=10)
        password_entry = tk.Entry(frame, font=("Arial", 12), width=25, show="●", relief="solid", bd=1)
        password_entry.grid(row=3, column=1, padx=10, pady=10)
        
        def register():
            if self.db.register_user(username_entry.get(), email_entry.get(), password_entry.get()):
                messagebox.showinfo("Success", "Registration successful!")
                self.show_login()
            else:
                messagebox.showerror("Error", "Username or email already exists!")
        
        tk.Button(frame, text="Register", font=("Arial", 12, "bold"), bg="#4CAF50", fg="white",
                 width=15, height=2, command=register, cursor="hand2").grid(row=4, column=0, columnspan=2, pady=20)
        
        tk.Button(frame, text="Back to Login", font=("Arial", 10), bg="#607D8B", fg="white",
                 command=self.show_login, cursor="hand2").grid(row=5, column=0, columnspan=2, pady=(0, 20))
    
    def show_home(self):
        self.clear_window()
        self.apply_theme()
        
        # Top Navigation Bar
        nav_frame = tk.Frame(self.root, bg=self.highlight, height=70)
        nav_frame.pack(fill="x")
        nav_frame.pack_propagate(False)
        
        tk.Label(nav_frame, text=f"👤 {self.current_user['username']}", 
                font=("Arial", 18, "bold"), bg=self.highlight, fg="white").pack(side="left", padx=20, pady=15)
        
        btn_frame = tk.Frame(nav_frame, bg=self.highlight)
        btn_frame.pack(side="right", padx=20)
        
        tk.Button(btn_frame, text="🌓 Toggle Theme", bg="#FFA726", fg="white", font=("Arial", 10), 
                 command=self.toggle_theme, cursor="hand2").pack(side="left", padx=5)
        
        tk.Button(btn_frame, text="⚙️ Budget", bg="#9C27B0", fg="white", font=("Arial", 10),
                 command=self.set_budget, cursor="hand2").pack(side="left", padx=5)
        
        tk.Button(btn_frame, text="🚪 Logout", bg="#f44336", fg="white", font=("Arial", 10),
                 command=self.show_login, cursor="hand2").pack(side="left", padx=5)
        
        # Main Container
        main_container = tk.Frame(self.root, bg=self.bg_color)
        main_container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Left Panel
        left_panel = tk.Frame(main_container, bg=self.bg_color, width=800)
        left_panel.pack(side="left", fill="both", expand=True, padx=(0, 5))
        
        # Summary Cards
        self.create_summary_cards(left_panel)
        
        # Date Filter
        self.create_date_filter(left_panel)
        
        # Add/Edit Expense
        self.create_expense_form(left_panel)
        
        # Expense List
        self.create_expense_list(left_panel)
        
        # Right Panel - Charts
        right_panel = tk.Frame(main_container, bg=self.bg_color, width=380)
        right_panel.pack(side="right", fill="both", padx=(5, 0))
        
        self.create_charts(right_panel)
        
        # Action Buttons
        self.create_action_buttons(left_panel)
    
    def create_summary_cards(self, parent):
        cards_frame = tk.Frame(parent, bg=self.bg_color)
        cards_frame.pack(fill="x", pady=(0, 10))
        
        # Total Expense Card
        card1 = tk.Frame(cards_frame, bg="#4CAF50", relief="raised", bd=2)
        card1.pack(side="left", fill="both", expand=True, padx=(0, 5))
        
        total = self.db.get_total_expense(self.current_user['id'])
        tk.Label(card1, text="💰 Total Expenses", font=("Arial", 12, "bold"), bg="#4CAF50", fg="white").pack(pady=(10, 5))
        tk.Label(card1, text=f"₹{total:.2f}", font=("Arial", 20, "bold"), bg="#4CAF50", fg="white").pack(pady=(0, 10))
        
        # Monthly Budget Card
        card2 = tk.Frame(cards_frame, bg="#FF9800", relief="raised", bd=2)
        card2.pack(side="left", fill="both", expand=True, padx=5)
        
        budget = self.current_user.get('budget', 0)
        remaining = budget - total if budget > 0 else 0
        
        tk.Label(card2, text="🎯 Budget Status", font=("Arial", 12, "bold"), bg="#FF9800", fg="white").pack(pady=(10, 5))
        
        if budget > 0:
            status = f"₹{remaining:.2f} Left" if remaining > 0 else f"₹{abs(remaining):.2f} Over!"
            tk.Label(card2, text=status, font=("Arial", 16, "bold"), bg="#FF9800", fg="white").pack(pady=(0, 10))
        else:
            tk.Label(card2, text="Not Set", font=("Arial", 16, "bold"), bg="#FF9800", fg="white").pack(pady=(0, 10))
        
        # This Month Card
        card3 = tk.Frame(cards_frame, bg="#2196F3", relief="raised", bd=2)
        card3.pack(side="left", fill="both", expand=True, padx=(5, 0))
        
        now = datetime.now()
        month_start = now.replace(day=1).strftime("%Y-%m-%d")
        month_total = self.db.get_total_expense(self.current_user['id'], month_start, now.strftime("%Y-%m-%d"))
        
        tk.Label(card3, text="📅 This Month", font=("Arial", 12, "bold"), bg="#2196F3", fg="white").pack(pady=(10, 5))
        tk.Label(card3, text=f"₹{month_total:.2f}", font=("Arial", 20, "bold"), bg="#2196F3", fg="white").pack(pady=(0, 10))
    
    def create_date_filter(self, parent):
        filter_frame = tk.LabelFrame(parent, text="📆 Date Range Filter", font=("Arial", 12, "bold"),
                                     bg=self.frame_bg, fg=self.fg_color, relief="raised", bd=2)
        filter_frame.pack(fill="x", pady=(0, 10))
        
        inner_frame = tk.Frame(filter_frame, bg=self.frame_bg)
        inner_frame.pack(pady=10, padx=10)
        
        tk.Label(inner_frame, text="From:", bg=self.frame_bg, fg=self.fg_color).grid(row=0, column=0, padx=5)
        self.start_date_var = tk.StringVar(value=(datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d"))
        start_entry = tk.Entry(inner_frame, textvariable=self.start_date_var, width=12)
        start_entry.grid(row=0, column=1, padx=5)
        
        tk.Label(inner_frame, text="To:", bg=self.frame_bg, fg=self.fg_color).grid(row=0, column=2, padx=5)
        self.end_date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        end_entry = tk.Entry(inner_frame, textvariable=self.end_date_var, width=12)
        end_entry.grid(row=0, column=3, padx=5)
        
        tk.Button(inner_frame, text="Apply Filter", bg="#2196F3", fg="white", command=self.apply_filter).grid(row=0, column=4, padx=5)
        tk.Button(inner_frame, text="Reset", bg="#607D8B", fg="white", command=self.reset_filter).grid(row=0, column=5, padx=5)
        
        # Quick Filters
        quick_frame = tk.Frame(filter_frame, bg=self.frame_bg)
        quick_frame.pack(pady=(0, 10))
        
        tk.Button(quick_frame, text="Today", bg="#4CAF50", fg="white", width=8, command=lambda: self.quick_filter('today')).pack(side="left", padx=3)
        tk.Button(quick_frame, text="This Week", bg="#4CAF50", fg="white", width=10, command=lambda: self.quick_filter('week')).pack(side="left", padx=3)
        tk.Button(quick_frame, text="This Month", bg="#4CAF50", fg="white", width=10, command=lambda: self.quick_filter('month')).pack(side="left", padx=3)
        tk.Button(quick_frame, text="This Year", bg="#4CAF50", fg="white", width=10, command=lambda: self.quick_filter('year')).pack(side="left", padx=3)
    
    def create_expense_form(self, parent):
        form_frame = tk.LabelFrame(parent, text="➕ Add / Edit Expense", font=("Arial", 12, "bold"),
                                   bg=self.frame_bg, fg=self.fg_color, relief="raised", bd=2)
        form_frame.pack(fill="x", pady=(0, 10))
        
        inner = tk.Frame(form_frame, bg=self.frame_bg)
        inner.pack(pady=10, padx=10)
        
        tk.Label(inner, text="Name:", bg=self.frame_bg, fg=self.fg_color).grid(row=0, column=0, padx=5, pady=5)
        self.name_var = tk.StringVar()
        tk.Entry(inner, textvariable=self.name_var, width=20).grid(row=0, column=1, padx=5, pady=5)
        
        tk.Label(inner, text="Category:", bg=self.frame_bg, fg=self.fg_color).grid(row=0, column=2, padx=5, pady=5)
        self.category_var = tk.StringVar(value="Other")
        ttk.Combobox(inner, textvariable=self.category_var, values=self.CATEGORIES, width=15, state="readonly").grid(row=0, column=3, padx=5, pady=5)
        
        tk.Label(inner, text="Amount:", bg=self.frame_bg, fg=self.fg_color).grid(row=0, column=4, padx=5, pady=5)
        self.amount_var = tk.StringVar()
        tk.Entry(inner, textvariable=self.amount_var, width=12).grid(row=0, column=5, padx=5, pady=5)
        
        tk.Label(inner, text="Date:", bg=self.frame_bg, fg=self.fg_color).grid(row=1, column=0, padx=5, pady=5)
        self.date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        tk.Entry(inner, textvariable=self.date_var, width=12).grid(row=1, column=1, padx=5, pady=5)
        
        tk.Label(inner, text="Time:", bg=self.frame_bg, fg=self.fg_color).grid(row=1, column=2, padx=5, pady=5)
        self.time_var = tk.StringVar(value=datetime.now().strftime("%H:%M"))
        tk.Entry(inner, textvariable=self.time_var, width=10).grid(row=1, column=3, padx=5, pady=5)
        
        self.edit_id = None
        
        tk.Button(inner, text="💾 Save", bg="#4CAF50", fg="white", width=10, command=self.save_expense).grid(row=1, column=4, padx=5, pady=5)
        tk.Button(inner, text="🔄 Clear", bg="#607D8B", fg="white", width=10, command=self.clear_form).grid(row=1, column=5, padx=5, pady=5)
    
    def create_expense_list(self, parent):
        list_frame = tk.LabelFrame(parent, text="📊 Expense List", font=("Arial", 12, "bold"),
                                   bg=self.frame_bg, fg=self.fg_color, relief="raised", bd=2)
        list_frame.pack(fill="both", expand=True, pady=(0, 10))
        
        # Scrollbar
        scroll = ttk.Scrollbar(list_frame)
        scroll.pack(side="right", fill="y")
        
        self.tree = ttk.Treeview(list_frame, columns=("ID", "Name", "Category", "Amount", "Date", "Time"),
                                show="headings", height=10, yscrollcommand=scroll.set)
        scroll.config(command=self.tree.yview)
        
        self.tree.heading("ID", text="ID")
        self.tree.heading("Name", text="Expense Name")
        self.tree.heading("Category", text="Category")
        self.tree.heading("Amount", text="Amount")
        self.tree.heading("Date", text="Date")
        self.tree.heading("Time", text="Time")
        
        self.tree.column("ID", width=50, anchor="center")
        self.tree.column("Name", width=200)
        self.tree.column("Category", width=120)
        self.tree.column("Amount", width=100, anchor="e")
        self.tree.column("Date", width=100, anchor="center")
        self.tree.column("Time", width=80, anchor="center")
        
        self.tree.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Buttons
        btn_frame = tk.Frame(list_frame, bg=self.frame_bg)
        btn_frame.pack(pady=(0, 10))
        
        tk.Button(btn_frame, text="✏️ Edit", bg="#FF9800", fg="white", width=10, command=self.edit_expense).pack(side="left", padx=5)
        tk.Button(btn_frame, text="🗑️ Delete", bg="#f44336", fg="white", width=10, command=self.delete_expense).pack(side="left", padx=5)
        
        self.load_expenses()
    
    def create_charts(self, parent):
        chart_frame = tk.LabelFrame(parent, text="📈 Expense Analytics", font=("Arial", 12, "bold"),
                                    bg=self.frame_bg, fg=self.fg_color, relief="raised", bd=2)
        chart_frame.pack(fill="both", expand=True)
        
        # Get category data
        cat_data = self.db.get_category_totals(self.current_user['id'])
        
        if cat_data:
            categories = [item[0] for item in cat_data]
            amounts = [float(item[1]) for item in cat_data]
            
            # Create pie chart
            fig, ax = plt.subplots(figsize=(4, 3.5), facecolor='none' if self.dark_mode else 'white')
            colors = ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF', '#FF9F40', '#FF6384', '#C9CBCF']
            
            ax.pie(amounts, labels=categories, autopct='%1.1f%%', startangle=90, colors=colors[:len(categories)])
            ax.set_title("Expenses by Category", fontsize=12, color=self.fg_color)
            
            canvas = FigureCanvasTkAgg(fig, chart_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(pady=10, padx=10)
        else:
            tk.Label(chart_frame, text="No data to display", font=("Arial", 14), 
                    bg=self.frame_bg, fg=self.fg_color).pack(expand=True)
    
    def create_action_buttons(self, parent):
        action_frame = tk.Frame(parent, bg=self.bg_color)
        action_frame.pack(fill="x")
        
        tk.Button(action_frame, text="📄 Export to PDF", bg="#9C27B0", fg="white", font=("Arial", 11, "bold"),
                 width=18, height=2, command=self.generate_pdf).pack(side="left", padx=(0, 5))
        
        tk.Button(action_frame, text="📊 Export to Excel", bg="#00796B", fg="white", font=("Arial", 11, "bold"),
                 width=18, height=2, command=self.export_to_excel).pack(side="left")
    
    def load_expenses(self, start_date=None, end_date=None):
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        expenses = self.db.get_expenses(self.current_user['id'], start_date, end_date)
        for exp in expenses:
            self.tree.insert("", "end", values=(exp[0], exp[1], exp[2], f"₹{exp[3]:.2f}", exp[4], exp[5]))
    
    def save_expense(self):
        try:
            if self.edit_id:
                self.db.update_expense(self.edit_id, self.name_var.get(), self.category_var.get(),
                                      float(self.amount_var.get()), self.date_var.get(), self.time_var.get())
                messagebox.showinfo("Success", "Expense updated!")
            else:
                self.db.add_expense(self.current_user['id'], self.name_var.get(), self.category_var.get(),
                                   float(self.amount_var.get()), self.date_var.get(), self.time_var.get())
                messagebox.showinfo("Success", "Expense added!")
            
            self.clear_form()
            self.show_home()
        except Exception as e:
            messagebox.showerror("Error", f"Failed: {str(e)}")
    
    def edit_expense(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select an expense to edit!")
            return
        
        item = self.tree.item(selected[0])
        values = item['values']
        
        self.edit_id = values[0]
        self.name_var.set(values[1])
        self.category_var.set(values[2])
        self.amount_var.set(str(values[3]).replace('₹', ''))
        self.date_var.set(values[4])
        self.time_var.set(values[5])
    
    def delete_expense(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select an expense to delete!")
            return
        
        if messagebox.askyesno("Confirm", "Are you sure you want to delete this expense?"):
            item = self.tree.item(selected[0])
            exp_id = item['values'][0]
            self.db.delete_expense(exp_id)
            messagebox.showinfo("Success", "Expense deleted!")
            self.show_home()
    
    def clear_form(self):
        self.edit_id = None
        self.name_var.set("")
        self.category_var.set("Other")
        self.amount_var.set("")
        self.date_var.set(datetime.now().strftime("%Y-%m-%d"))
        self.time_var.set(datetime.now().strftime("%H:%M"))
    
    def apply_filter(self):
        start = self.start_date_var.get()
        end = self.end_date_var.get()
        self.load_expenses(start, end)
    
    def reset_filter(self):
        self.start_date_var.set((datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d"))
        self.end_date_var.set(datetime.now().strftime("%Y-%m-%d"))
        self.load_expenses()
    
    def quick_filter(self, period):
        now = datetime.now()
        
        if period == 'today':
            start = end = now.strftime("%Y-%m-%d")
        elif period == 'week':
            start = (now - timedelta(days=now.weekday())).strftime("%Y-%m-%d")
            end = now.strftime("%Y-%m-%d")
        elif period == 'month':
            start = now.replace(day=1).strftime("%Y-%m-%d")
            end = now.strftime("%Y-%m-%d")
        elif period == 'year':
            start = now.replace(month=1, day=1).strftime("%Y-%m-%d")
            end = now.strftime("%Y-%m-%d")
        
        self.start_date_var.set(start)
        self.end_date_var.set(end)
        self.load_expenses(start, end)
    
    def set_budget(self):
        budget = tk.simpledialog.askfloat("Set Budget", "Enter monthly budget (₹):", 
                                         initialvalue=self.current_user.get('budget', 0))
        if budget is not None:
            self.db.update_budget(self.current_user['id'], budget)
            self.current_user['budget'] = budget
            messagebox.showinfo("Success", f"Budget set to ₹{budget:.2f}")
            self.show_home()
    
    def toggle_theme(self):
        self.dark_mode = not self.dark_mode
        self.db.toggle_dark_mode(self.current_user['id'], self.dark_mode)
        self.show_home()
    
    def generate_pdf(self):
        expenses = self.db.get_expenses(self.current_user['id'])
        
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", "B", 18)
        pdf.cell(0, 15, f"Expense Report - {self.current_user['username']}", ln=True, align="C")
        pdf.ln(5)
        
        pdf.set_font("Arial", "", 10)
        pdf.cell(0, 10, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ln=True)
        pdf.ln(5)
        
        pdf.set_font("Arial", "B", 11)
        pdf.cell(50, 10, "Name", border=1)
        pdf.cell(35, 10, "Category", border=1)
        pdf.cell(30, 10, "Amount", border=1)
        pdf.cell(35, 10, "Date", border=1)
        pdf.cell(30, 10, "Time", border=1, ln=True)
        
        pdf.set_font("Arial", "", 9)
        total = 0
        for exp in expenses:
            pdf.cell(50, 10, str(exp[1])[:20], border=1)
            pdf.cell(35, 10, str(exp[2])[:15], border=1)
            pdf.cell(30, 10, f"Rs {exp[3]:.2f}", border=1)
            pdf.cell(35, 10, str(exp[4]), border=1)
            pdf.cell(30, 10, str(exp[5]), border=1, ln=True)
            total += exp[3]
        
        pdf.ln(5)
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 10, f"Total Expenses: Rs {total:.2f}", ln=True)
        
        filename = f"expense_report_{self.current_user['username']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        pdf.output(filename)
        messagebox.showinfo("Success", f"PDF saved as {filename}")
    
    def export_to_excel(self):
        expenses = self.db.get_expenses(self.current_user['id'])
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"
        
        # Headers
        headers = ["ID", "Expense Name", "Category", "Amount (₹)", "Date", "Time"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data
        total = 0
        for exp in expenses:
            ws.append([exp[0], exp[1], exp[2], float(exp[3]), str(exp[4]), str(exp[5])])
            total += exp[3]
        
        # Total row
        ws.append(["", "", "TOTAL", total, "", ""])
        last_row = ws.max_row
        ws[f"C{last_row}"].font = Font(bold=True)
        ws[f"D{last_row}"].font = Font(bold=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        
        filename = f"expense_report_{self.current_user['username']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        messagebox.showinfo("Success", f"Excel file saved as {filename}")
    
    def show_admin_dashboard(self):
        self.clear_window()
        self.root.configure(bg="#f0f0f0")
        
        header = tk.Frame(self.root, bg="#673AB7", height=80)
        header.pack(fill="x")
        tk.Label(header, text="🔧 Admin Dashboard", font=("Arial", 22, "bold"), 
                bg="#673AB7", fg="white").pack(pady=20)
        
        tk.Button(header, text="Logout", bg="#f44336", fg="white", font=("Arial", 11),
                 command=self.show_login).place(relx=0.95, rely=0.3, anchor="e")
        
        list_frame = tk.LabelFrame(self.root, text="👥 Registered Users", font=("Arial", 14, "bold"), 
                                   bg="#ffffff", padx=20, pady=10)
        list_frame.pack(pady=20, padx=20, fill="both", expand=True)
        
        tree = ttk.Treeview(list_frame, columns=("Username", "Email", "Created", "Status"), 
                           show="headings", height=18)
        tree.heading("Username", text="Username")
        tree.heading("Email", text="Email")
        tree.heading("Created", text="Registration Date")
        tree.heading("Status", text="Status")
        
        tree.column("Username", width=200)
        tree.column("Email", width=300)
        tree.column("Created", width=200)
        tree.column("Status", width=150)
        
        users = self.db.get_all_users()
        for user in users:
            status = "🟢 Active" if user[3] else "🔴 Inactive"
            tree.insert("", "end", values=(user[0], user[1], user[2], status))
        
        tree.pack(fill="both", expand=True)
        
        tk.Label(list_frame, text=f"Total Users: {len(users)}", font=("Arial", 13, "bold"), 
                bg="#ffffff").pack(pady=15)

if __name__ == "__main__":
    root = tk.Tk()
    app = ExpenseTrackerApp(root)
    root.mainloop()